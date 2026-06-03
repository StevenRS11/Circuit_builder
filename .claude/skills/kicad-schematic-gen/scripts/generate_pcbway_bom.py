#!/usr/bin/env python3
"""
PCBway BOM generator (Stage 9) — transform the internal Stage-3 BOM markdown into
PCBway's upload .xlsx form.

PCBway's turnkey quote form has a fixed 16-column layout (we have a real example
from a returned quote). The customer fills the first nine columns; PCBway fills
the rest (price, delivery, actual purchase part, notes, etc.). This script is a
pure, deterministic *transform* — it makes no part decisions. It:

  - parses the internal BOM (reusing check_pcbway.load_bom_for_pcbway),
  - groups identical parts onto one line (comma-joined designators + qty),
  - maps our columns onto PCBway's, and
  - writes the .xlsx with the customer columns filled and PCBway's columns blank.

Because it only relabels/relays existing data, the generated form is exactly as
good as the source BOM — which is the point: the Stage-3 bom.md is the single
source of truth, and this regenerates the upload form from it every time
(no hand-editing the spreadsheet, which is how stale/wrong cells crept in before).

The internal BOM must carry a **Manufacturer** column and keep distributor codes
(LCSC C…) OUT of the Part Number column — see templates/03_bom.md.

CLI:
    python generate_pcbway_bom.py <bom.md>                       # -> <project>_PCBway_BOM.xlsx beside the BOM
    python generate_pcbway_bom.py <bom.md> -o out.xlsx
    python generate_pcbway_bom.py <bom.md> --output-dir PCBway_uploads
    python generate_pcbway_bom.py <bom.md> --json
"""

import sys
import os
import re
import json as json_module

# Reuse the canonical BOM parser so there is exactly one BOM format to maintain.
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from check_pcbway import load_bom_for_pcbway  # noqa: E402


# PCBway's official BOM-upload template columns, in order (matches their sample
# "a-sample-of-PCBWay-BOM.xlsx"). Every line needs a real Manufacturer + Mfg Part #
# (passives included). There is no distributor-PN column — the LCSC/DigiKey code
# goes in the Notes column, which is how PCBway picks it up for sourcing.
PCBWAY_COLUMNS = [
    "Item #", "*Designator", "*Qty", "Manufacturer", "*Mfg Part #",
    "Description / Value", "*Package/Footprint", "Type", "Your Instructions / Notes",
]


# Reference-designator prefix -> description type word.
_TYPE_WORDS = [
    ("RT", "THERMISTOR"), ("RS", "RES"), ("FB", "FERRITE"),
    ("R", "RES"), ("C", "CAP"), ("L", "IND"),
    ("U", "IC"), ("Q", "TRANSISTOR"), ("D", "DIODE"),
    ("J", "CONN"), ("CN", "CONN"), ("SW", "SWITCH"), ("Y", "XTAL"), ("X", "XTAL"),
]


def _type_word(reference, footprint):
    if footprint and re.search(r"LED", footprint, re.IGNORECASE):
        return "LED"
    for prefix, word in _TYPE_WORDS:
        if reference.upper().startswith(prefix):
            return word
    return ""


def describe(part):
    """'Description / Value' string.

    Prefer the BOM's explicit Description column (rich, e.g. 'CAP CER 47uF 25V
    X5R 1210'); fall back to a synthesized '{TYPE} {value} {package}' when the
    BOM has no Description for the line.
    """
    if getattr(part, "description", "").strip():
        return part.description.strip()
    word = _type_word(part.reference, part.footprint)
    pkg = part.package or ""
    bits = [b for b in (word, part.value, pkg) if b]
    return " ".join(dict.fromkeys(bits)).strip()


# ─── helpers ─────────────────────────────────────────────────────────

_THT_MARKERS = re.compile(
    r"_THT|PinHeader|PinSocket|TerminalBlock|TO-?220|TO-?247|TO-?92|TO-?263|"
    r"DIP-|_DIP\b|_P2\.54mm|Through", re.IGNORECASE)


def mounting_type(footprint):
    """SMD by default; THT only on a clear through-hole marker.

    Note: a JST '..._Horizontal' SMD side-entry connector is still SMD, so we do
    NOT treat 'Horizontal'/'Vertical' as THT here (that would mislabel SMD connectors).
    """
    return "THT" if footprint and _THT_MARKERS.search(footprint) else "SMD"


def _natural_key(ref):
    """Sort key so R2 < R10 < R100 (prefix alpha, then numeric)."""
    m = re.match(r"^([A-Za-z]+)(\d*)", ref.strip())
    if not m:
        return (ref, 0)
    return (m.group(1), int(m.group(2)) if m.group(2) else 0)


def _blank(s):
    """Treat placeholder dashes (—, -, –) and empty as no-value."""
    s = (s or "").strip()
    return "" if s in ("", "-", "—", "–", "n/a", "N/A", "TBD") else s


def _vendor_part(part):
    """'LCSC C970725' style vendor reference from supplier + supplier_pn."""
    sup, pn = _blank(part.supplier), _blank(part.supplier_pn)
    if sup and pn:
        return f"{sup} {pn}"
    return pn or sup


def group_parts(parts):
    """Group identical parts onto one PCBway line.

    Identity = (manufacturer, MPN, value, package, footprint, supplier, supplier_pn).
    Per-instance Notes/function are intentionally excluded from the key (PCBway
    groups by part, not by role). Group order follows first appearance in the BOM.
    """
    groups = {}
    order = []
    for p in parts:
        key = (p.manufacturer, p.part_number, p.value, p.package,
               p.footprint, p.supplier, p.supplier_pn)
        if key not in groups:
            groups[key] = {"refs": [], "part": p}
            order.append(key)
        groups[key]["refs"].append(p.reference)

    lines = []
    for key in order:
        g = groups[key]
        p = g["part"]
        refs = sorted(g["refs"], key=_natural_key)
        lines.append({
            "designators": ",".join(refs),
            "qty": len(refs),
            "manufacturer": p.manufacturer,
            "mfg_part": p.part_number,
            "description": describe(p),
            "package": p.package or p.footprint,
            "type": mounting_type(p.footprint),
            "notes": _vendor_part(p),   # LCSC/DigiKey code → PCBway sources from it
        })
    return lines


def _project_title(md_text, bom_path):
    """Descriptive name for the in-sheet title (may include a parenthetical)."""
    m = re.search(r"#\s*Bill of Materials\s*[—-]\s*(.+)", md_text)
    if m:
        return m.group(1).strip()
    return _project_slug(bom_path)


def _project_slug(bom_path):
    """Filename-safe project slug from the BOM filename (strips the _NN_bom suffix)."""
    stem = os.path.splitext(os.path.basename(bom_path))[0]
    stem = re.sub(r"_\d*_?bom(_flat)?$", "", stem, flags=re.IGNORECASE)
    return stem or "project"


# ─── xlsx writer ─────────────────────────────────────────────────────

def write_xlsx(lines, out_path, project_name):
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BOM"

    title = f"{project_name} — PCBA BOM (PCBway turnkey)"
    ws.cell(1, 1, title).font = Font(bold=True, size=12)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(PCBWAY_COLUMNS))

    hdr_row = 2
    hdr_fill = PatternFill("solid", fgColor="D9E1F2")
    for c, name in enumerate(PCBWAY_COLUMNS, start=1):
        cell = ws.cell(hdr_row, c, name)
        cell.font = Font(bold=True)
        cell.fill = hdr_fill
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    for i, ln in enumerate(lines):
        r = hdr_row + 1 + i
        vals = [
            i + 1, ln["designators"], ln["qty"], ln["manufacturer"], ln["mfg_part"],
            ln["description"], ln["package"], ln["type"], ln["notes"],
        ]
        for c, v in enumerate(vals, start=1):
            ws.cell(r, c, v).alignment = Alignment(wrap_text=True, vertical="top")

    widths = [7, 24, 6, 18, 24, 30, 18, 8, 22]
    for c, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = w

    os.makedirs(os.path.dirname(os.path.abspath(out_path)), exist_ok=True)
    wb.save(out_path)


# ─── orchestration ───────────────────────────────────────────────────

def generate(bom_path, out_path=None, output_dir=None):
    with open(bom_path, "r", encoding="utf-8") as f:
        md = f.read()
    parts = load_bom_for_pcbway(md)
    if not parts:
        raise ValueError(f"No BOM rows parsed from {bom_path} — is it a Stage-3 BOM table?")
    title = _project_title(md, bom_path)
    slug = _project_slug(bom_path)
    lines = group_parts(parts)

    if not out_path:
        fname = f"{slug}_PCBway_BOM.xlsx"
        base = output_dir or os.path.dirname(os.path.abspath(bom_path))
        out_path = os.path.join(base, fname)

    write_xlsx(lines, out_path, title)
    return {
        "bom": os.path.abspath(bom_path),
        "output": os.path.abspath(out_path),
        "project": slug,
        "part_count": len(parts),
        "line_count": len(lines),
    }


def main():
    import argparse
    parser = argparse.ArgumentParser(
        description="Generate the PCBway upload .xlsx from a Stage-3 BOM markdown.")
    parser.add_argument("bom", help="Stage-3 BOM markdown file")
    parser.add_argument("-o", "--output", help="Output .xlsx path")
    parser.add_argument("--output-dir", help="Directory for the default-named .xlsx")
    parser.add_argument("--json", action="store_true", help="JSON output")
    args = parser.parse_args()

    try:
        report = generate(args.bom, out_path=args.output, output_dir=args.output_dir)
    except (FileNotFoundError, ValueError) as e:
        print(f"ERROR: {e}", file=sys.stderr)
        sys.exit(1)

    if args.json:
        print(json_module.dumps(report, indent=2))
    else:
        print(f"Wrote {report['output']}")
        print(f"  {report['part_count']} parts -> {report['line_count']} BOM lines")
    sys.exit(0)


if __name__ == "__main__":
    main()
