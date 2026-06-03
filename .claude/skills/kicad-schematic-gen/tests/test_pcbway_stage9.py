#!/usr/bin/env python3
"""Tests for the Stage 9 PCBway-upload pipeline:
  - check_pcbway.py structural line gates (distributor-code-in-MPN, package mismatch,
    description-not-MPN, missing manufacturer, verification_worklist)
  - generate_pcbway_bom.py (grouping, 9-col transform, every line has an MPN)
  - bom_verify.py (worklist + verdict aggregation / mismatch gate)
  - generate_fab_outputs.py (layer-stack detection, kicad-cli locator)
"""

import sys
import os

_tests_dir = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, os.path.abspath(os.path.join(_tests_dir, "..", "scripts")))

import pytest
from check_pcbway import (
    check_bom, PcbwayPart, load_bom_for_pcbway,
    looks_like_distributor_code, looks_like_description,
    footprint_size_token, package_size_token, verification_worklist,
)
import generate_pcbway_bom as gp
import bom_verify as bv
import generate_fab_outputs as gf


def _p(ref, **kw):
    return PcbwayPart(reference=ref, **kw)


def _checks(result):
    return {i.check_name for i in result.issues}


# A small but realistic Stage-3 BOM (post-schema: Manufacturer + Description + MPN).
GOOD_BOM = """# Bill of Materials — demo_board

| Ref | Value | Manufacturer | Part Number | Description | Package | KiCad Symbol (lib_id) | Footprint (KiCad) | In Stock? | Qty | Supplier | Supplier PN | Unit Price | Notes |
|-----|-------|--------------|-------------|-------------|---------|-----------------------|-------------------|-----------|-----|----------|-------------|------------|-------|
| U1 | RegIC | Diodes Inc | AP2112K-3.3 | IC LDO 3.3V | SOT-23-5 | Regulator_Linear:AP2112K-3.3 | Package_TO_SOT_SMD:SOT-23-5 | yes | 1 | LCSC | C51118 | $0.10 | LDO |
| R1 | 10k | YAGEO | RC0805FR-0710KL | RES 10k 1% 0805 | 0805 | Device:R | Resistor_SMD:R_0805_2012Metric | yes | 1 | LCSC | C84376 | $0.01 | pull-up |
| R2 | 10k | YAGEO | RC0805FR-0710KL | RES 10k 1% 0805 | 0805 | Device:R | Resistor_SMD:R_0805_2012Metric | yes | 1 | LCSC | C84376 | $0.01 | pull-up |
| C1 | 100nF | Samsung | CL10B104KB8NNNC | CAP CER 100nF 50V 0402 | 0402 | Device:C | Capacitor_SMD:C_0402_1005Metric | yes | 1 | LCSC | C1525 | $0.01 | decouple |
"""


# ─── check_pcbway structural helpers ─────────────────────────────────

class TestStructuralHelpers:
    def test_distributor_code_detection(self):
        assert looks_like_distributor_code("C914291")
        assert looks_like_distributor_code("C13564")
        assert not looks_like_distributor_code("NCP15XH103F03RC")  # real MPN
        assert not looks_like_distributor_code("CL21A475KAQNNNE")  # cap MPN starting with C
        assert not looks_like_distributor_code("")

    def test_description_detection(self):
        assert looks_like_description("56k 5% 0805")
        assert not looks_like_description("RC0805FR-0756KL")
        assert not looks_like_description("")

    def test_size_tokens(self):
        assert footprint_size_token("Resistor_SMD:R_0805_2012Metric") == "0805"
        assert footprint_size_token("LED_SMD:LED_0603_1608Metric") == "0603"
        assert footprint_size_token("Package_TO_SOT_SMD:SOT-23-6") is None
        assert package_size_token("0805") == "0805"
        assert package_size_token("SOT-23-6") is None


# ─── the field-defect catches ────────────────────────────────────────

class TestStructuralGate:
    def test_distributor_code_in_mpn_blocks(self):
        # the C914291=Zener class: a valid LCSC code in the MPN column
        r = check_bom([_p("RS1", part_number="C914291", manufacturer="X",
                          package="1206", footprint="Resistor_SMD:R_1206_3216Metric")])
        assert not r.passed
        assert "distributor_code_as_mpn" in _checks(r)

    def test_package_footprint_mismatch_blocks(self):
        # the RT1 class: package field says 0805, footprint says 0402
        r = check_bom([_p("RT1", part_number="NCP15XH103F03RC", manufacturer="Murata",
                          package="0805", footprint="Resistor_SMD:R_0402_1005Metric")])
        assert not r.passed
        assert "package_mismatch" in _checks(r)

    def test_description_in_mpn_warns(self):
        r = check_bom([_p("R1", value="56k", part_number="56k 5% 0805", manufacturer="YAGEO",
                          package="0805", footprint="Resistor_SMD:R_0805_2012Metric")])
        assert "mpn_not_real" in _checks(r)

    def test_missing_manufacturer_warns(self):
        r = check_bom([_p("D7", value="SS34", part_number="SS34", manufacturer="",
                          package="SMA", footprint="Diode_SMD:D_SMA")])
        assert "missing_manufacturer" in _checks(r)

    def test_clean_bom_passes(self):
        r = check_bom(load_bom_for_pcbway(GOOD_BOM))
        assert r.passed
        assert "distributor_code_as_mpn" not in _checks(r)
        assert "package_mismatch" not in _checks(r)
        assert "missing_manufacturer" not in _checks(r)

    def test_parser_reads_manufacturer_and_description(self):
        parts = {p.reference: p for p in load_bom_for_pcbway(GOOD_BOM)}
        assert parts["U1"].manufacturer == "Diodes Inc"
        assert parts["U1"].description == "IC LDO 3.3V"


class TestWorklist:
    def test_worklist_non_passives_plus_flagged(self):
        parts = load_bom_for_pcbway(GOOD_BOM)
        check_bom(parts)
        refs = {p.reference for p in verification_worklist(parts)}
        assert "U1" in refs          # non-passive always
        assert "R1" not in refs      # clean passive skipped
        assert "C1" not in refs

    def test_worklist_all_includes_passives(self):
        parts = load_bom_for_pcbway(GOOD_BOM)
        check_bom(parts)
        refs = {p.reference for p in verification_worklist(parts, include_passives=True)}
        assert {"U1", "R1", "C1"} <= refs


# ─── generate_pcbway_bom ─────────────────────────────────────────────

class TestPcbwayBomGenerator:
    def test_grouping_and_qty(self):
        lines = gp.group_parts(load_bom_for_pcbway(GOOD_BOM))
        by_mpn = {ln["mfg_part"]: ln for ln in lines}
        # R1 + R2 (identical) collapse to one line, qty 2, designators joined
        r = by_mpn["RC0805FR-0710KL"]
        assert r["qty"] == 2
        assert r["designators"] == "R1,R2"

    def test_every_line_has_mpn_and_manufacturer(self):
        lines = gp.group_parts(load_bom_for_pcbway(GOOD_BOM))
        assert all(ln["mfg_part"] for ln in lines)
        assert all(ln["manufacturer"] for ln in lines)

    def test_description_prefers_bom_column(self):
        parts = load_bom_for_pcbway(GOOD_BOM)
        u1 = next(p for p in parts if p.reference == "U1")
        assert gp.describe(u1) == "IC LDO 3.3V"

    def test_description_synthesized_when_absent(self):
        p = _p("R9", value="4.7R", package="0805", footprint="Resistor_SMD:R_0805_2012Metric")
        assert gp.describe(p) == "RES 4.7R 0805"

    def test_mounting_type(self):
        assert gp.mounting_type("Resistor_SMD:R_0805_2012Metric") == "SMD"
        assert gp.mounting_type("Connector_PinHeader_2.54mm:PinHeader_1x02_P2.54mm_Vertical") == "THT"
        # an SMD side-entry connector is NOT THT despite "_Horizontal"
        assert gp.mounting_type("Connector_JST:JST_PH_S2B-PH-SM4-TB_1x02-1MP_P2.00mm_Horizontal") == "SMD"

    def test_vendor_part_drops_placeholder_dash(self):
        assert gp._vendor_part(_p("R1", supplier="LCSC", supplier_pn="—")) == "LCSC"
        assert gp._vendor_part(_p("R1", supplier="LCSC", supplier_pn="C84376")) == "LCSC C84376"

    def test_project_slug_strips_bom_suffix(self):
        assert gp._project_slug("battery_3s_03_bom.md") == "battery_3s"

    def test_generate_writes_9col_xlsx(self, tmp_path):
        pytest.importorskip("openpyxl")
        import openpyxl
        bom = tmp_path / "demo_03_bom.md"
        bom.write_text(GOOD_BOM, encoding="utf-8")
        rep = gp.generate(str(bom), output_dir=str(tmp_path))
        assert rep["line_count"] == 3   # R1+R2 grouped
        ws = openpyxl.load_workbook(rep["output"])["BOM"]
        assert ws.cell(2, 1).value == "Item #"
        assert ws.max_column == len(gp.PCBWAY_COLUMNS) == 9


# ─── bom_verify ──────────────────────────────────────────────────────

class TestBomVerify:
    def test_worklist_claims(self):
        items = bv.build_worklist(load_bom_for_pcbway(GOOD_BOM))
        u1 = next(it for it in items if it["ref"] == "U1")
        assert u1["claim"]["mpn"] == "AP2112K-3.3"
        assert u1["claim"]["distributor_code"] == "LCSC C51118"

    def test_report_flags_mismatch_and_counts(self):
        parts = load_bom_for_pcbway(GOOD_BOM)
        verdicts = [
            {"ref": "U1", "verdict": "mismatch", "package_match": False,
             "distributor_resolves": False, "lifecycle": "active",
             "evidence": "code resolves to a different part"},
        ]
        md, n_mismatch = bv.build_report(parts, verdicts)
        assert n_mismatch == 1
        assert "MISMATCH" in md

    def test_report_clean_when_confirmed(self):
        parts = load_bom_for_pcbway(GOOD_BOM)
        verdicts = [{"ref": "U1", "verdict": "confirmed", "package_match": True,
                     "distributor_resolves": True, "lifecycle": "active", "evidence": "ok"}]
        _md, n_mismatch = bv.build_report(parts, verdicts)
        assert n_mismatch == 0


# ─── generate_fab_outputs ────────────────────────────────────────────

class TestFabOutputs:
    def test_detect_4layer_stack(self):
        pcb = '''(kicad_pcb (layers
          (0 "F.Cu" signal)
          (1 "GND" signal)
          (2 "GND2" signal)
          (31 "B.Cu" signal)
          (32 "B.Adhes" user)
        )
        )'''
        assert gf.detect_copper_layers(pcb) == ["F.Cu", "In1.Cu", "In2.Cu", "B.Cu"]

    def test_detect_2layer_stack(self):
        pcb = '(kicad_pcb (layers (0 "F.Cu" signal) (31 "B.Cu" signal)\n))'
        assert gf.detect_copper_layers(pcb) == ["F.Cu", "B.Cu"]

    def test_kicad_cli_locator_bad_override_raises(self):
        with pytest.raises(FileNotFoundError):
            gf.find_kicad_cli(override="/nonexistent/path/to/kicad-cli")
